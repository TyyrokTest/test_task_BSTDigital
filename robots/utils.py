from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font

SHEET_TITLE = ["Модель", "Версия", "Количество за неделю"]

def createExcel(qs):
    """Writing data to Excel in temporary file
    Args:
        qs (QuerySet): Contains 'model', 'version', 'num' 
    Returns:
        stream tempfile: excel tmp file in stream
    """
    wb = Workbook()  
    ws = None
    ft = Font(bold=True)
    
    for robot in qs:

        if ws is None:
            ws = wb.active
            ws.title = robot['model']
            prev_model = robot['model']
            ws.append(SHEET_TITLE)
            for row in ws['A1:C1']:
                for cell in row:
                    cell.font = ft
                    
        elif prev_model != robot['model']:
            prev_model = robot['model']
            ws = wb.create_sheet(prev_model)
            ws.append(SHEET_TITLE)
            for row in ws['A1:C1']:
                for cell in row:
                    cell.font = ft
            
        data_row = []
        data_row.append(robot['model'])
        data_row.append(robot['version']) 
        data_row.append(robot['num'])
        ws.append(data_row)
        
    with NamedTemporaryFile() as tmp:
        wb.save(tmp)
        tmp.seek(0)
        stream = tmp.read()
        
    return stream